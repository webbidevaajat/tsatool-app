#!/usr/bin/python
# -*- coding: utf-8 -*-

# Collection of CondCollections

import logging
import os
import psycopg2
import openpyxl as xl
from .cond_collection import CondCollection
from .error import TsaErrCollection
from .utils import trunc_str
from .utils import list_local_statids
from .utils import list_local_sensors
from datetime import datetime
from collections import OrderedDict

DEFAULT_PG_HOST = 'localhost'
DEFAULT_PG_PORT = 5432
DEFAULT_PG_DBNAME = 'tsa'
DEFAULT_PG_USER = 'postgres'
DEFAULT_PG_PASSWORD = 'postgres'

PPTX_TEMPLATE_PATH = 'report_template.pptx'

log = logging.getLogger(__name__)

class DBParams:
    """
    Stores parameters for database connection.
    """
    def __init__(self):
        self.dbname = os.getenv('PG_DBNAME', DEFAULT_PG_DBNAME)
        self.user = os.getenv('PG_USER', DEFAULT_PG_USER)
        self.password = os.getenv('PG_PASSWORD', DEFAULT_PG_PASSWORD)
        self.host = os.getenv('PG_HOST', DEFAULT_PG_HOST)
        self.port = os.getenv('PG_PORT', DEFAULT_PG_PORT)

    def keys(self):
        return ['dbname', 'user', 'password', 'host', 'port']

    def __getitem__(self, key):
        return self.__dict__[key]

    def __str__(self):
        s = 'DBParams\n'
        for k in self.keys():
            if k == 'password':
                v = '(not shown)'
            else:
                v = self.__dict__[k]
            s += f'{k:8}: {v}, '
        return s

class AnalysisCollection:
    """
    A collection of ``CondCollection`` instances.
    Can be used to run multiple analyses as a batch job.
    Enables validating CondCollections separately
    and then analysing them.

    Output files are saved in ``results/`` relative to current working directory.
    PowerPoint reports follow pattern ``results/[name][_sheetname].pptx``,
    and Excel files ``results/[name].xlsx``.
    Existing output files with same filepath will be overwritten.
    """
    def __init__(self, input_xlsx, name):
        self.created_at = datetime.now()
        self.input_xlsx = input_xlsx
        self.name = name
        self.workbook = xl.load_workbook(filename=input_xlsx, read_only=True)

        os.makedirs('results', exist_ok=True)
        self.out_base_path = f'results/{self.name}'

        # This will contain condition rows by Excel sheet,
        # read by a separate method
        self.collections = OrderedDict()

        # Station ids and sensor name-id pairs for dryvalidate checking.
        # These represent the situation as of 8/2019.
        # Updates must be made to corresponding functions in tsa.utils.
        self.local_statids = list_local_statids()
        self.local_sensor_pairs = list_local_sensors()

        # DB connection is made by a separate method only if needed;
        # dryvalidate methods are available also without it.
        self.db_params = DBParams()
        self.db_statids = set()
        self.db_sensor_pairs = dict()

        # Errors are reported on the fly AND collected too
        self.errors = TsaErrCollection('ANALYSIS / EXCEL FILE')

    def add_collections(self, drop=['info']):
        """
        Add CondCollections from worksheets.
        :param drop: list of sheets to exclude by title
        :type drop: list of strings
        """
        sheetnames = [s for s in self.workbook.sheetnames if s.lower().strip() not in drop]
        for title in sheetnames:
            try:
                self.collections[title] = CondCollection.from_xlsx_sheet(
                    ws=self.workbook[title]
                )
                log.info(f'Added CondCollection <{title}>')
            except:
                self.errors.add(msg=f'Could not add CondCollection <{title}>: skipping',
                                log_add='exception')

    def set_sensor_ids(self, pairs):
        """
        Set sensor name-id pairs for all ``Blocks``.

        :param pairs: dict, key = sensor id, value = sensor name
        """
        for coll in self.collections.keys():
            for cnd in self[coll].conditions.keys():
                for bl in self[coll][cnd].blocks.keys():
                    self[coll][cnd][bl].set_sensor_id(pairs)

    def validate_statids_with_set(self, station_ids):
        """
        For all primary ``Blocks``, check if their station ids are valid,
        i.e., they can be found in ``station_ids`` set.
        """
        station_ids = set(station_ids)
        for d in self.collections.keys():
            for c in self.collections[d].conditions.keys():
                for b in self.collections[d].conditions[c].blocks.keys():
                    isprimary = self.collections[d].conditions[c].blocks[b].secondary is False
                    hasid = self.collections[d].conditions[c].blocks[b].station_id is not None
                    validstatid = self.collections[d].conditions[c].blocks[b].station_id in station_ids
                    if not isprimary:
                        continue
                    if not hasid:
                        self.collections[d].conditions[c].blocks[b].errors.add(
                            msg='stationid is None (tried to compare it to static ids)',
                            log_add='error'
                        )
                        continue
                    if not validstatid:
                        self.collections[d].conditions[c].blocks[b].errors.add(
                            msg='stationid was not found in static ids',
                            log_add='error'
                        )

    def collect_errors(self):
        """
        Collect error messages from all levels
        into an OrderedDict tree.

        :return: tuple ``(True if any errors, OrderedDict of errors)``
        """
        log.debug('Building error message tree ...')
        haserrs = False
        master = OrderedDict(
            errors = [str(e) for e in self.errors.errors],
            collections = OrderedDict()
        )
        if master['errors']:
            haserrs = True
        for coll in self.collections.values():
            colldict = OrderedDict(
                errors = [str(e) for e in coll.errors.errors],
                conditions = OrderedDict()
            )
            if colldict['errors']:
                haserrs = True
            for cond in coll.conditions.values():
                conddict = OrderedDict(
                    errors = [str(e) for e in cond.errors.errors],
                    blocks = OrderedDict()
                )
                if conddict['errors']:
                    haserrs = True
                for block in cond.blocks.values():
                    blockdict = OrderedDict(
                        errors = [str(e) for e in block.errors.errors]
                    )
                    if blockdict['errors']:
                        haserrs = True
                    conddict['blocks'][str(block)] = blockdict
                colldict['conditions'][str(cond)] = conddict
            master['collections'][str(coll)] = colldict
        return haserrs, master

    def run_analyses(self):
        """
        Run analyses for CondCollections that were made from the selected Excel sheets,
        and save results according to the selected formats and path names.
        Analyses are run against collection-specific db connections.
        """
        log.debug(f'Initializing Excel workbook for {str(self)}')
        wb = xl.Workbook()
        ws = wb.active
        ws.title = 'INFO'
        ws['A1'].value = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        ws['B1'].value = 'analysis started'
        wb_path = f'{self.out_base_path}_report.xlsx'
        log.debug(f'Excel workbook will be saved as {wb_path}')

        # Prepare directory for png images for pptx;
        # keep the pngs if png_dir is passed to
        # passed to CondCollection.to_pptx().
        png_dir = f'{self.out_base_path}_images'
        os.makedirs(png_dir, exist_ok=True)
        log.debug(f'Png images will be saved to {png_dir}')

        for cl in self.collections.keys():
            try:
                with psycopg2.connect(**self.db_params) as pg_conn:
                    coll_pptx_path = f'{self.out_base_path}_{cl}.pptx'
                    self.collections[cl].run_analysis(pg_conn=pg_conn,
                                                      wb=wb,
                                                      wb_path=wb_path,
                                                      pptx_path=coll_pptx_path,
                                                      pptx_template=PPTX_TEMPLATE_PATH,
                                                      png_dir=png_dir)
                    log.debug(f'{str(self.collections[cl])} is analyzed')
            except:
                self.errors.add(
                    msg=f'Skipping {str(self.collections[cl])} due to fatal error',
                    log_add='exception'
                )

        wb['INFO']['A2'].value = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        wb['INFO']['B2'].value = 'analysis ended'
        wb.save(wb_path)
        log.debug(f'Excel workbook saved as {wb_path}')
        log.info(f'{str(self)} analyzed')

    def __getitem__(self, key):
        """
        Return the CondCollection from the OrderedDict referenced by ``key``.
        """
        return self.collections[key]

    def __str__(self):
        s = f'<AnalysisCollection {self.name}> from <{self.input_xlsx}> '
        s += f'with {len(self.collections)} collections>'
        return s
