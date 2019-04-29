#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
This module contains tools for operating with road weather station database
using condition strings and aliases.

This module shall be imported by Dash ``app.py``.

Note that ``db_config.json`` file must exist in the root directory
to connect to the TSA database.
"""
import os
import re
import json
import pandas
import psycopg2
import traceback
import openpyxl as xl
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from matplotlib import rcParams
from getpass import getpass
from datetime import datetime
from datetime import timedelta

# Set matplotlib parameters globally
rcParams['font.family'] = 'sans-serif'
rcParams['font.sans-serif'] = ['Arial', 'Tahoma']

def tsadb_connect(username=None, password=None, ask=False):
    """
    Using ``db_config.json`` file in the root directory,
    connect to the TSA database.
    :param username: database username as string; defaults to
                     'ADMIN_USER' of the config file
    :param password: database user password, only for debugging!
                     Defaults to None -> will be asked
    :param ask:      if ``True``, use interactive input for username;
                     defaults to ``False``
    :return:         connection instance, or None on error
    """
    cf_filename = os.path.join(os.getcwd(), 'db_config.json')
    with open(cf_filename, 'r') as cf_file:
        cf = json.load(cf_file)
    if username is None:
        if ask:
            username = input('Database username: ')
        else:
            username = cf['ADMIN_USER']
    if password is None:
        password = getpass('Password for user "{:s}": '.format(username))
    try:
        pg_conn = psycopg2.connect(dbname=cf['DATABASE'],
                                   user=username,
                                   password=password,
                                   host=cf['HOST'],
                                   port=cf['PORT'],
                                   connect_timeout=5)
        return pg_conn
    except psycopg2.OperationalError as e:
        print('Could not connect to database:')
        print(e)
        print('Are you connected to the right network?')
        print('Using correct and existing username and password?')
        return None

def eliminate_umlauts(x):
    """
    Converts ä and ö into a and o.
    """
    umlauts = {
        'ä': 'a',
        'Ä': 'A',
        'ö': 'o',
        'Ö': 'O'
    }
    for k in umlauts.keys():
        x = x.replace(k, umlauts[k])

    return x

def to_pg_identifier(x):
    """
    Converts x (string) such that it can be used as a table or column
    identifier in PostgreSQL.

    If there are whitespaces in the middle,
    they are converted into underscores.

    Raises error if x contains fatally invalid parts, e.g.
    leading digit or a non-alphanumeric character.

    .. note:: Pg identifier length maximum is 63 characters.
        To avoid too long final identifiers
        (that might be concatenated from multiple original ones),
        max length of x here
        is 40 characters, which should be enough for site names too.
    """
    x = x.strip()

    # Original string without leading/trailing whitespaces
    # is retained for error prompting purposes
    old_x = x
    x = x.lower()
    x = eliminate_umlauts(x)
    x = x.replace(' ', '_')

    if x[0].isdigit():
        errtext = 'String starts with digit:\n'
        errtext += old_x + '\n'
        errtext += '^'
        raise ValueError(errtext)

    if len(x) > 40:
        errtext = 'String too long, maximum is 40 characters:\n'
        errtext += old_x + '\n'
        raise ValueError(errtext)

    for i, c in enumerate(x):
        if not (c.isalnum() or c == '_'):
            errtext = 'String contains an invalid character:\n'
            errtext += old_x + '\n'
            errtext += '~' * i + '^'
            raise ValueError(errtext)

    return x

def strfdelta(tdelta, fmt):
    """
    Format timedelta object according to ``fmt`` string.
    ``fmt`` should contain formatting string with placeholders
    ``{days}``, ``{hours}``, ``{minutes}`` and ``{seconds}``.
    """
    d = {'days': tdelta.days}
    d['hours'], rem = divmod(tdelta.seconds, 3600)
    d['minutes'], d['seconds'] = divmod(rem, 60)
    return fmt.format(**d)

class Block:
    """
    Represents a logical subcondition
    with information of site name and station id.
    See :py:class:`Condition` that consists of blocks
    and operators and parentheses.
    A Block renders as boolean column in temporary db tables.
    For PostgreSQL compatibility, umlauts convert to a and o,
    and all strings are made lowercase.

    For a *primary* block, the `raw_condition` must consist of
    a station identifier, hashtag, sensor identifier,
    operator and a value.
    For a *secondary* block, the `raw_condition` must consist of
    a site identifier, hashtag and alias identifier. Note that these should
    refer to an existing Condition instance.

    :example::

        # Making a primary block:
        >>> Block('d2', 'ylojarvi_etelaan_2', 3, 's1122#kitka3_luku >= 0.30')
        {'raw_logic': 's1122#kitka3_luku >= 0.30',
        'master_alias': 'd2',
        'parent_site': 'ylojarvi_etelaan_2',
        'alias': 'd2_3',
        'secondary': False,
        'site': 'ylojarvi_etelaan_2',
        'station': 's1122',
        'source_alias': None,
        'sensor': 'kitka3_luku',
        'operator': '>=',
        'value_str': '0.30'}
        # Making a secondary block:
        >>> Block('d2', 4, 'ylojarvi_pohjoiseen_1#c3')
        {'ylojarvi_pohjoiseen_1#c3',
        'master_alias': 'd2',
        'parent_site': 'ylojarvi_etelaan_2',
        'alias': 'd2_4',
        'secondary': True,
        'site': 'ylojarvi_pohjoiseen_1',
        'station': None,
        'source_alias': 'c3',
        'sensor': None,
        'operator': None,
        'value_str': None}

    :param master_alias: master alias identifier of the parent condition
    :type master_alias: string
    :param parent_site: site identifier of the parent condition
    :type parent_site: string
    :param order_nr: index of the block within the parent condition
    :type order_nr: integer
    :param raw_logic: logic to parse, bound to single sensor or existing Condition
    :type raw_logic: string
    """
    def __init__(self, master_alias, parent_site, order_nr, raw_logic):
        self.raw_logic = raw_logic
        self.master_alias = to_pg_identifier(master_alias)
        self.parent_site = to_pg_identifier(parent_site)
        self.order_nr = order_nr
        self.alias = self.master_alias + '_' + str(order_nr)
        self.secondary = None
        self.site = None
        self.station = None
        self.station_id = None
        self.source_alias = None
        self.sensor = None
        self.sensor_id = None
        self.operator = None
        self.value_str = None

        # Set values depending on raw logic given
        self.unpack_logic()

    def error_context(self, before='', after=''):
        """
        Return context information on block,
        to be used with error messages.
        """
        s = before
        s += f'\nBlock {self.alias}:\n'
        s += after
        return s

    def unpack_logic(self):
        """
        Detects and sets block type and attributes from raw logic string
        and checks validity of the attributes.

        .. note:: Following binary operators are considered:
            `'=', '<>', '>', '<', '>=', '<=', 'in'`.
            `between` is currently not supported.
            If operator is `in`, it is checked whether the value after it
            is a valid SQL tuple.
            Operator MUST be surrounded by whitespaces!

        :param raw_logic: original logic string
        :type raw_logic: string
        """
        binops = [' = ', ' <> ', ' > ', ' < ', ' >= ', ' <= ', ' in ']

        # ERROR if too many hashtags or operators
        n_hashtags = self.raw_logic.count('#')
        if n_hashtags > 1:
            errtext = 'Too many "#"s, only one or zero allowed:\n'
            errtext = self.error_context(after=errtext)
            raise ValueError(errtext)
        n_binops = 0
        binop_in_str = None
        for binop in binops:
            if binop in self.raw_logic:
                n_binops += self.raw_logic.count(binop)
                binop_in_str = binop
        if n_binops > 1:
            errtext = 'Too many binary operators, only one or zero allowed:\n'
            errtext = self.error_context(after=errtext)
            raise ValueError(errtext)

        # Case 1: contains no hashtag and no binary operator
        # -> secondary block, site is picked from parent_site.
        # Must NOT contain binary operator.
        if n_hashtags == 0 and n_binops == 0:
            self.secondary = True
            self.site = self.parent_site
            try:
                self.source_alias = to_pg_identifier(self.raw_logic)
            except ValueError as e:
                raise ValueError(self.error_context(after=e))

        # Case 2: contains hashtag but no binary operator
        # -> secondary block
        elif n_hashtags == 1 and n_binops == 0:
            self.secondary = True
            parts = self.raw_logic.split('#')
            try:
                self.site = to_pg_identifier(parts[0])
                self.source_alias = to_pg_identifier(parts[1])
            except ValueError as e:
                raise ValueError(self.error_context(after=e))

        # Case 3: contains hashtag and binary operator
        # -> primary block
        elif n_hashtags == 1 and n_binops == 1:
            self.secondary = False
            self.site = self.parent_site
            parts = self.raw_logic.split('#')
            parts = [parts[0]] + parts[1].split(binop_in_str)
            try:
                self.station = to_pg_identifier(parts[0])
                self.station_id = int(''.join(i for i in self.station if i.isdigit()))
                self.sensor = to_pg_identifier(parts[1])
                self.operator = binop_in_str.lower().strip()
                self.value_str = parts[2].lower().strip()
            except ValueError as e:
                raise ValueError(self.error_context(after=e))

            # Special case with operator "in":
            # must be followed by tuple enclosed with parentheses.
            if self.operator == 'in':
                val_sw = self.value_str.startswith('(')
                val_ew = self.value_str.endswith(')')
                if val_sw is False and val_ew is False:
                    errtext = 'Binary operator "in" must be followed by\n'
                    errtext += 'a tuple enclosed with parentheses "()":\n'
                    errtext += self.raw_logic
                    raise ValueError(self.error_context(after=errtext))

        # Case 4: ERROR if binary operator but no hashtag
        else:
            errtext = 'No "#" given, should be of format\n'
            errtext += '[station]#[sensor] [binary operator] [value]:\n'
            errtext += self.raw_logic
            raise ValueError(self.error_context(after=errtext))

    def set_sensor_id(self, nameids):
        """
        Set sensor id based on name-id dict,
        presumably gotten from database.
        """
        try:
            self.sensor_id = nameids[self.sensor]
        except KeyError:
            errtext = f"Sensor '{self.sensor}' not found in database."
            raise KeyError(self.error_context(after=errtext))

    def get_sql_def(self):
        """
        Create SQL ``pack_ranges`` function call
        string to be used as part of the corresponding
        Condition table creation.
        """
        if not self.sensor_id:
            errtext = 'No sensor_id set for\n'
            errtext += str(self)
            raise Exception(self.error_context(after=errtext))

        if self.secondary:
            errtext = 'Analyzing secondary blocks is not yet supported:\n'
            errtext += str(self)
            raise Exception(self.error_context(after=errtext))

        sql = (f"SELECT valid_r, istrue AS {self.alias} "
               "FROM pack_ranges("
               "p_obs_relation := 'obs_main', "
               "p_maxminutes := 30, "
               f"p_statid := {self.station_id}, "
               f"p_seid := {self.sensor_id}, "
               f"p_operator := '{self.operator}', "
               f"p_seval := '{self.value_str}')")

        return sql

    def __str__(self):
        if self.secondary:
            s = 'Secondary '
        else:
            s = 'Primary '
        s += 'Block {:s} at {:s}: {:s}'.format(
            self.alias, self.parent_site, self.raw_logic)
        return s

    def __repr__(self):
        # TODO: need to make representation more unambiguous
        # compared to __str__? Current one might be a bad workaround.
        return str(self)

    def __eq__(self, other):
        """
        The `==` method; two blocks are equal if their attributes
        are equal, **including the order number in ** `self.alias`.
        """
        return self.__dict__ == other.__dict__


class Condition:
    """
    Single condition, its aliases, query handling and results.

    :example::

        # Making a primary condition:
        >>> Condition('Ylöjärvi_etelään_1',
        ... 'E4',
        ... '(s1122#TIENPINNAN_TILA3 = 8 OR (s1122#KITKA3_LUKU >= 0.30
        ... AND s1122#KITKA3_LUKU < 0.4)) AND s1115#TIE_1 < 2',
        ... ('2018-01-01 00:00', '2018-04-30 23:59'))
        Primary Condition ylojarvi_etelaan_1_e4:
        (s1122#tienpinnan_tila3 = 8 or (s1122#kitka3_luku >= 0.30 and
        s1122#kitka3_luku < 0.4)) and s1115#tie_1 < 2
        ALIAS: (e4_0 or (e4_1 and e4_2)) and e4_3
        >>>
        # Making a secondary condition:
        >>> Condition('Ylöjärvi_pohjoiseen_2',
        ... 'B1',
        ... 'D2 AND D3',
        ... ('2018-01-01 00:00', '2018-04-30 23:59'))
        Secondary Condition ylojarvi_pohjoiseen_2_b1:
        d2 and d3
        ALIAS: b1_0 and b1_1
        >>>
        # Making a secondary condition containing primary blocks:
        >>> Condition('Ylöjärvi_etelään_2',
        ... 'C3',
        ... 'Ylöjärvi_pohjoiseen_2#B1 AND s1011#TIE_1 > 0',
        ... ('2018-01-01 0:00', '2018-04-30 23:59'))
        Secondary Condition ylojarvi_etelaan_2_c3:
        ylojarvi_pohjoiseen_2#b1 and s1011#tie_1 > 0
        ALIAS: c3_0 and c3_1

    :param site: site / location / area identifier
    :type site: string
    :param master_alias: master alias identifier
    :type master_alias: string
    :param raw_condition: condition definition
    :type raw_condition: string
    :param time_range: start (included) and end (included) timestamps
    :type time_range: list or tuple of datetime objects
    """
    def __init__(self, site, master_alias, raw_condition, time_range, excel_row=None):
        # Excel row for prompting, if made from Excel sheet
        self.excel_row = excel_row

        # List for saving error strings
        self.errmsgs = []

        # Original formattings are kept for printing purposes
        self.orig_site = site
        self.orig_master_alias = master_alias
        self.orig_condition = raw_condition

        # Attrs for further use must be PostgreSQL compatible
        self.site = to_pg_identifier(site)
        self.master_alias = to_pg_identifier(master_alias)
        self.id_string = '{:s}_{:s}'.format(self.site, self.master_alias)

        # TODO: wrap condition str handling to function or property/setter??
        raw_condition = raw_condition.strip().lower()
        raw_condition = eliminate_umlauts(raw_condition)
        self.condition = raw_condition

        # Times must be datetime objects
        self.time_from = time_range[0]
        self.time_until = time_range[1]

        # As above but representing actual min and max time of result data
        self.data_from = None
        self.data_until = None

        # make_blocks() sets .blocks, .alias_condition and .secondary
        self.blocks = None
        self.alias_condition = None
        self.secondary = None
        self.make_blocks()

        self.stations = set()
        self.list_stations()

        self.has_view = False

        # pandas DataFrames for results
        self.main_df = None
        self.durations_df = None

        # Total time will be set to represent
        # actual min and max timestamps of the data
        self.tottime = self.time_until - self.time_from
        self.tottime_valid = None
        self.tottime_notvalid = None
        self.tottime_nodata = None
        self.percentage_valid = None
        self.percentage_notvalid = None
        self.percentage_nodata = None

    def error_context(self, before='', after=''):
        """
        Return information on condition and its Excel row if available,
        to be used with error messages.
        """
        s = before + '\n'
        s += f'ERROR with condition {self.id_string}'
        if self.excel_row:
            s += f'\n(row {self.excel_row} in Excel sheet):\n'
        else:
            s += ':\n' + after
        return s

    def make_blocks(self):
        """
        Extract a list of Block instances (that is, subconditions)
        into `self.blocks` based on `self.condition`,
        define `self.alias_condition` based on the aliases of the Block instances
        and detect condition type (`secondary == True` if any of the blocks has
        `secondary == True`, `False` otherwise).
        """
        value = self.condition

        # Generally, opening and closing bracket counts must match
        n_open = value.count('(')
        n_close = value.count(')')
        if n_open != n_close:
            errtext = self.error_context()
            errtext += 'Unequal number of opening and closing parentheses:\n'
            errtext += '{:d} opening and {:d} closing'.format(n_open, n_close)
            self.errmsgs.append(errtext)
            raise ValueError(errtext)

        # Eliminate multiple whitespaces
        # and leading and trailing whitespaces
        value = ' '.join(value.split()).strip()

        # Split by
        # - parentheses
        # - and, or, not: must be surrounded by spaces
        # - not: starting the string and followed by space.
        # Then strip results from trailing and leading whitespaces
        # and remove empty elements.
        sp = re.split('([()]|(?<=\s)and(?=\s)|(?<=\s)or(?=\s)|(?<=\s)not(?=\s)|^not(?=\s))', value)
        sp = [el.strip() for el in sp]
        sp = [el for el in sp if el]

        # Handle special case of parentheses after "in":
        # they are part of the logic element.
        # Block() will detect in the next step
        # if the tuple after "in" is not correctly enclosed by ")".
        new_sp = []
        for el in sp:
            if not new_sp:
                new_sp.append(el)
                continue
            if len(new_sp[-1]) > 3 and new_sp[-1][-3:] == ' in':
                new_sp[-1] = new_sp[-1] + ' ' + el
            elif ' in ' in new_sp[-1] and new_sp[-1][-1] != ')':
                new_sp[-1] = new_sp[-1] + el
            else:
                new_sp.append(el)

        # Identify the "role" of each element by making them into
        # tuples like (role, element).
        # First, mark parentheses and and-or-not operators.
        # The rest should convert to logic blocks;
        # Block() raises error if this does not succeed.
        idfied = []
        i = 0
        for el in new_sp:
            if el == '(':
                idfied.append(('open_par', el))
            elif el == ')':
                idfied.append(('close_par', el))
            elif el in ['and', 'or']:
                idfied.append(('andor', el))
            elif el == 'not':
                idfied.append(('not', el))
            else:
                bl = Block(master_alias=self.master_alias,
                    parent_site=self.site,
                    order_nr=i,
                    raw_logic=el)
                # If a block with same contents already exists,
                # do not add a new one with another order number i,
                # but add the existing block with its order number.
                # The .index() method raises an error in case the tuple with
                # Block element is NOT contained in the list.
                existing_blocks = [t for t in idfied if t[0] == 'block']
                for eb in existing_blocks:
                    if eb[1].raw_logic == bl.raw_logic:
                        idfied.append(eb)
                        break
                else:
                    idfied.append(('block', bl))
                    i += 1

        def validate_order(tuples):
            """
            Validate order of the elements of a :py:class:`Block`.

            :param tuples: list of tuples, each of which has
                `open_par`, `close_par`, `andor`, `not` or `block`
                in the first index and the string element itself in the second.
            :type tuples: list or tuple
            :return: no return if element order is valid, otherwise
                raise an error upon first invalid element

            Following element types may be in the first index:
                `open_par`, `not`, `block`

            Following elements may be in the last index:
                `close_par`, `block`

            For elements other than the last one, see the table below
            to see what element can follow each element.
            Take the first element from left and the next element from top.

            +-------------+------------+-------------+---------+-------+---------+
            |             | `open_par` | `close_par` | `andor` | `not` | `block` |
            +=============+============+=============+=========+=======+=========+
            | `open_par`  | OK         | X           | X       | OK    | OK      |
            +-------------+------------+-------------+---------+-------+---------+
            | `close_par` | X          | OK          | OK      | X     | X       |
            +-------------+------------+-------------+---------+-------+---------+
            | `andor`     | OK         | X           | X       | OK    | OK      |
            +-------------+------------+-------------+---------+-------+---------+
            | `not`       | OK         | X           | X       | X     | OK      |
            +-------------+------------+-------------+---------+-------+---------+
            | `block`     | X          | OK          | OK      | X     | X       |
            +-------------+------------+-------------+---------+-------+---------+
            """
            allowed_first = ('open_par', 'not', 'block')
            allowed_pairs = (
            ('open_par', 'open_par'), ('open_par', 'not'), ('open_par', 'block'),
            ('close_par', 'close_par'), ('close_par', 'andor'),
            ('andor', 'open_par'), ('andor', 'not'), ('andor', 'block'),
            ('not', 'open_par'), ('not', 'block'),
            ('block', 'close_par'), ('block', 'andor')
            )
            allowed_last = ('close_par', 'block')
            last_i = len(tuples) - 1

            for i, el in enumerate(tuples):
                if i == 0:
                    if el[0] not in allowed_first:
                        errtext = '{"{:s}" not allowed as first element:\n'.format(el[1])
                        errtext = self.error_context(after=errtext)
                        self.errmsgs.append(errtext)
                        raise ValueError(errtext)
                elif i == last_i:
                    if el[0] not in allowed_last:
                        errtext = '"{:s}" not allowed as last element:\n'.format(el[1])
                        errtext = self.error_context(after=errtext)
                        self.errmsgs.append(errtext)
                        raise ValueError(errtext)
                if i < last_i:
                    if (el[0], tuples[i+1][0]) not in allowed_pairs:
                        errtext = '"{:s}" not allowed right before "{:s}":\n'.format(el[1], tuples[i+1][1])
                        errtext = self.error_context(after=errtext)
                        self.errmsgs.append(errtext)
                        raise ValueError(errtext)

        # Check the correct order of the tuples.
        # This should raise and error and thus exit the method
        # if there is an illegal combination of elements next to each other.
        validate_order(idfied)

        # If validation was successful, attributes can be set

        # Pick up all unique blocks in the order they appear
        blocks = []
        for el in idfied:
            if el[0] == 'block' and el[1] not in blocks:
                blocks.append(el[1])
        self.blocks = sorted(blocks, key=lambda x: x.alias)

        # Form alias condition by replacing original block parts by
        # the alias of the corresponding block
        al_value = value
        for bl in self.blocks:
            al_value = al_value.replace(bl.raw_logic, bl.alias)
        self.alias_condition = al_value

        # If any of the blocks is secondary,
        # then the whole condition is considered secondary.
        self.secondary = False
        for bl in self.blocks:
            if bl.secondary:
                self.secondary = True
                break

    def list_stations(self):
        """
        Add unique stations of primary `self.blocks`
        to `self.stations` set
        """
        for bl in self.blocks:
            if not bl.secondary:
                self.stations.add(bl.station)

    def create_db_view(self, pg_conn=None, verbose=False):
        """
        Create temporary table corresponding to the condition.
        Effective only if a database connection is present.
        """
        # TODO: define secondary condition behaviour

        drop_sql = f"DROP VIEW IF EXISTS {self.id_string};\n"
        sql = f"CREATE OR REPLACE TEMP VIEW {self.id_string} AS ( \nWITH "

        bl_defs = [f"{bl.alias} AS ({bl.get_sql_def()})" for bl in self.blocks]
        sql += ', \n'.join(bl_defs)

        if len(self.blocks) < 2:
            last_cte = self.blocks[0].alias
        else:
            sql += ', \n'

            # Agglomerate time range join CTEs
            ctes = []
            i = 2
            while i <= len(self.blocks):
                subsel = self.blocks[-i:]
                nr_seq = '_'.join([str(bl.order_nr) for bl in subsel])
                idfier = f"{self.master_alias}_{nr_seq}"
                if not ctes:
                    last_cte = subsel[-1].alias
                else:
                    last_cte = ctes[-1][0]
                cte_def = f"{idfier} AS (SELECT \n"
                cte_def += f"{subsel[0].alias}.valid_r * {last_cte}.valid_r AS valid_r, \n"
                cte_def += ', \n'.join([bl.alias for bl in subsel]) + ' \n'
                cte_def += f"FROM {subsel[0].alias} \n"
                cte_def += f"JOIN {last_cte} \n"
                cte_def += f"ON {subsel[0].alias}.valid_r && {last_cte}.valid_r)"
                ctes.append((idfier, cte_def))
                i += 1

            last_cte = ctes[-1][0]
            sql += ', \n'.join([cte[1] for cte in ctes])

        sql += ' \n'
        sql += ("SELECT \n"
                "lower(valid_r) AS vfrom, \n"
                "upper(valid_r) AS vuntil, \n"
                "upper(valid_r)-lower(valid_r) AS vdiff, \n")
        sql += ", \n".join([bl.alias for bl in self.blocks]) + ", \n"
        sql += f"({self.alias_condition}) AS master \n"
        sql += f"FROM {last_cte});"

        if verbose:
            print(drop_sql)
            print(sql)

        if not pg_conn:
            return

        with pg_conn.cursor() as cur:
            try:
                cur.execute(drop_sql)
                cur.execute(sql)
                pg_conn.commit()
                self.has_view = True
            except psycopg2.DatabaseError as e:
                pg_conn.rollback()
                errtext = self.error_context(after=e)
                self.errmsgs.append(errtext)
                raise psycopg2.DatabaseError(errtext)

    def set_summary_attrs(self):
        """
        Calculate summary attribute values using the ``.main_df`` DataFrame.
        """
        if self.main_df is None:
            return
        df = self.main_df

        self.tottime_valid = df[df['master']==True]['vdiff'].sum() or timedelta(0)
        self.tottime_notvalid = df[df['master']==False]['vdiff'].sum() or timedelta(0)
        self.tottime_nodata = self.tottime - self.tottime_valid - self.tottime_notvalid
        tts = self.tottime.total_seconds()
        self.percentage_valid = self.tottime_valid.total_seconds() / tts
        self.percentage_notvalid = self.tottime_notvalid.total_seconds() / tts
        self.percentage_nodata = self.tottime_nodata.total_seconds() / tts

    def calculate_durations_df(self):
        """
        Set a DataFrame of how long each condition is valid / not valid / null
        at a time.
        """
        # TODO: implement this later
        pass


    def fetch_results_from_db(self, pg_conn=None):
        """
        Fetch result data from corresponding db view
        to pandas DataFrame, and set summary attribute values
        based on the DataFrame.
        """
        if not pg_conn:
            return
        if not self.has_view:
            print('Could not fetch results for')
            print(f'{str(self)}:\n')
            print('since it does not have a corresponding database view.')
            return
        sql = f"SELECT * FROM {self.id_string};"
        self.main_df = pandas.read_sql(sql, con=pg_conn)

        self.data_from = self.main_df['vfrom'].min()
        self.data_until = self.main_df['vuntil'].max()
        if not (self.data_from is None or self.data_until is None):
            self.tottime = self.data_until - self.data_from

        self.set_summary_attrs()

    def get_timelineplot(self):
        """
        Returns a Matplotlib figure object:
        a `broken_barh` plot of the validity of the condition
        and its blocks on a timeline.
        """
        if self.main_df is None:
            raise Exception('No data to visualize.')

        def getfacecolor(val):
            """
            Return a color name
            by boolean column value.
            """
            if val == True:
                return '#f03b20'
            elif val == False:
                return '#2b83ba'
            return '#bababa'

        # Set height and transparency for block rows, between 0-1;
        # master row will be set to height 0.8 and alpha 1 below.
        hgtval = 0.5
        alphaval = 0.5
        # Offset of the logic label above the bar
        lbl_offset = 0.1

        # Make matplotlib-ready range list from the time columns
        xr = zip([mdates.date2num(el) for el in self.main_df['vfrom']],
                 [mdates.date2num(el) for el in self.main_df['vuntil']])
        xr = [(a, b-a) for (a, b) in xr]

        # Make subplots for blocks;
        # for every block, there should be
        # a corresponding boolean column in the result DataFrame!
        fig, ax = plt.subplots()
        yticks = []
        ylabels = []
        i = 1
        for bl in self.blocks:
            logic_lbl = bl.raw_logic
            ax.broken_barh(xranges=xr, yrange=(i, hgtval),
                           facecolors=list(map(getfacecolor,
                                               self.main_df[bl.alias])),
                           alpha=alphaval)
            ax.annotate(s=logic_lbl,
                        xy=(xr[0][0], i + hgtval + lbl_offset))
            yticks.append(i + (hgtval / 2))
            ylabels.append(bl.alias)
            i += 1

        # Add master row to the plot
        hgtval = 0.8
        ax.broken_barh(xranges=xr, yrange=(i, hgtval),
                       facecolors=list(map(getfacecolor,
                                           self.main_df['master'])))
        ax.annotate(s=self.alias_condition,
                    xy=(xr[0][0], i + hgtval + lbl_offset))
        yticks.append(i + (hgtval / 2))
        ylabels.append('master')
        i += 1

        # Set a whole lot of axis parameters...
        ax.set_axisbelow(True)

        ax.xaxis_date()
        ax.xaxis.set_major_formatter(mdates.DateFormatter('%m/%y'))
        ax.xaxis.set_major_locator(mdates.MonthLocator())
        ax.xaxis.set_ticks_position('none')
        ax.xaxis.grid(color='#e5e5e5')
        #plt.xticks(rotation=45)

        ax.set_yticks(yticks)
        ax.set_yticklabels(ylabels)
        ax.yaxis.set_ticks_position('none')

        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)
        ax.spines['left'].set_visible(False)
        ax.spines['bottom'].set_visible(False)

        return ax

    def __getitem__(self, key):
        """
        Returns the Block instance on the corresponding index.
        `key` can be integer or `.alias` string.
        """
        try:
            idx = int(key)
        except ValueError:
            idx = None
            for i, bl in enumerate(self.blocks):
                if bl.alias == key:
                    idx = i
                    break
            if idx is None:
                raise KeyError(f"No Block with alias '{key}'")
        return self.blocks[idx]


    def __str__(self):
        if self.secondary:
            s = '  Secondary '
        else:
            s = '  Primary '
        s += 'Condition {:s}:\n'.format(self.id_string)
        s += '    {:s}\n'.format(self.condition)
        s += '    ALIAS: {:s}'.format(self.alias_condition)
        return s

    def __repr__(self):
        # TODO unambiguous representation?
        return str(self)

class CondCollection:
    """
    A collection of conditions to analyze.
    All conditions share the same analysis time range.
    Times are assumed to be given as dates only,
    and their HH:MM:SS are overridden to default values,
    see ``set_default_times(self)``.

    # TODO CondCollection init parameters and results

    :param time_from: start time (inclusive) of the analysis period
    :type time_from: Python ``datetime()`` object
    :param time_until: end time (exclusive) of the analysis period
    :type time_until: Python ``datetime()`` object
    :param pg_conn: database connection
    :type pg_conn: ``psycopg2.connect()`` object
    """
    def __init__(self, time_from, time_until, pg_conn=None, title=None):
        # Times must be datetime objects and in correct order
        assert isinstance(time_from, datetime)
        assert isinstance(time_until, datetime)
        assert time_from < time_until
        self.time_from = time_from
        self.time_until = time_until
        self.set_default_times()
        self.time_range = (self.time_from, self.time_until)

        self.title = title

        # Timestamp is based on instance creation time,
        # not on when the analysis has been run
        self.created_timestamp = datetime.now()

        self.conditions = []
        self.stations = set()
        self.id_strings = set()

        self.pg_conn = pg_conn

        self.setup_statobs_view()
        self.statids_available = self.get_stations_in_view()

        self.setup_obs_view()

    def set_default_times(self):
        """
        Sets analysis start time to 00:00:00
        and end time to 23:59:59 on selected dates, respectively.
        """
        self.time_from = self.time_from.replace(
            hour=0, minute=0, second=0)
        self.time_until = self.time_until.replace(
            hour=23, minute=59, second=59)

    def setup_statobs_view(self, verbose=False):
        """
        In the database, create or replace a temporary view ``statobs_time``
        containing the station observations within the ``time_range``.
        """
        if self.pg_conn:
            with self.pg_conn.cursor() as cur:
                sql = ("CREATE OR REPLACE TEMP VIEW statobs_time AS "
                       "SELECT id, tfrom, statid "
                       "FROM statobs "
                       "WHERE tfrom BETWEEN %s AND %s;")
                if verbose:
                    print(cur.mogrify(sql, (self.time_from, self.time_until)))
                try:
                    cur.execute(sql, (self.time_from, self.time_until))
                    self.pg_conn.commit()
                except:
                    self.pg_conn.rollback()
                    print(traceback.print_exc())
        else:
            if verbose:
                print('No db connection, cannot create view')

    def get_stations_in_view(self):
        """
        Get stations available in ``statobs_time`` view.
        """
        if self.pg_conn:
            with self.pg_conn.cursor() as cur:
                sql = "SELECT DISTINCT statid FROM statobs_time ORDER BY statid;"
                cur.execute(sql, (self.time_from, self.time_until))
                statids = cur.fetchall()
                statids = [el[0] for el in statids]
                return statids
        else:
            return None

    def setup_obs_view(self):
        """
        After creating the ``statobs_time`` view,
        create a joint temporary view ``obs_main``
        that works as the main source for Block queries.
        """
        if not self.pg_conn:
            return
        with self.pg_conn.cursor() as cur:
            sql = ("CREATE OR REPLACE TEMP VIEW obs_main AS "
                   "SELECT tfrom, statid, seid, seval "
                   "FROM statobs_time "
                   "INNER JOIN seobs "
                   "ON statobs_time.id = seobs.obsid;")
            try:
                cur.execute(sql)
                self.pg_conn.commit()
            except:
                self.pg_conn.rollback()
                print(traceback.print_exc())

    def add_station(self, station):
        """
        Add ``station`` to ``self.stations`` if not already there.
        If db connection is available, check if the main view
        contains that station id.

        .. note::   station identifier must contain station integer id
                    when letters are removed.
        """
        if station not in self.stations:
            stid = int(''.join(i for i in station if i.isdigit()))
            if self.pg_conn:
                if stid not in self.statids_available:
                    print(f'WARNING: no observations for station {stid} in database!')
            self.stations.add(station)

    def add_condition(self, site, master_alias, raw_condition, excel_row=None):
        """
        Add new Condition instance, raise error if one exists already
        with same site-master_alias identifier.
        """
        candidate = Condition(site, master_alias, raw_condition, self.time_range, excel_row)
        if candidate.id_string in self.id_strings:
            errtext = 'Identifier {:s} is already reserved, cannot add\n'.format(candidate.id_string)
            errtext += raw_condition
            raise ValueError(errtext)
        else:
            self.conditions.append(candidate)
            self.id_strings.add(candidate.id_string)
            for s in candidate.stations:
                self.add_station(s)

    def set_sensor_ids(self):
        """
        Get sensor name - id correspondence from the database,
        and set sensor ids for all Blocks in all Conditions.
        """
        if not self.pg_conn:
            return
            # TODO: action upon no pg connection?
        with self.pg_conn.cursor() as cur:
            cur.execute("SELECT id, lower(name) AS name FROM sensors;")
            tb = cur.fetchall()
            nameids = {k:v for v, k in tb}
        for cnd in self.conditions:
            for bl in cnd.blocks:
                bl.set_sensor_id(nameids)

    def create_condition_views(self, verbose=False):
        """
        For each Condition, try to create the corresponding
        database view.
        """
        # TODO: secondary Condition handling,
        #       make primary ones first and
        #       then try secondary ones until they
        #       find existing identifiers from the database
        #       to create their views.
        for cnd in self.conditions:
            if cnd.secondary:
                print(f'{str(cnd)}\n')
                print('Secondary conditions not yet supported.')
                continue
            try:
                cnd.create_db_view(pg_conn=self.pg_conn, verbose=verbose)
            except Exception as e:
                print(f'{str(cnd)}:\n')
                print('Could not create a view.')
                print(traceback.print_exc())

    def fetch_all_results(self):
        """
        Fetch results
        for all Conditions that have a corresponding view in the database.
        """
        for cnd in self.conditions:
            try:
                cnd.fetch_results_from_db(pg_conn=self.pg_conn)
            except Exception as e:
                print(f'{str(cnd)}:\n')
                print('Could not fetch results from database.')
                print(traceback.print_exc())

    def to_workbook(self):
        """
        Return an ``openpyxl.Workbook`` with a single worksheet
        containing basic data of the condition collection.
        """
        wb = xl.Workbook()
        ws = wb.active
        ws.title = self.title or 'conditions'

        # Headers in fixed cells & styling
        headers = {'A1': 'start',
                   'B1': 'end',
                   'D1': 'analyzed',
                   'A3': 'site',
                   'B3': 'master_alias',
                   'C3': 'condition',
                   'D3': 'data_from',
                   'E3': 'data_until',
                   'F3': 'valid',
                   'G3': 'notvalid',
                   'H3': 'nodata'
                   }
        for k, v in headers.items():
            ws[k] = v
            ws[k].font = xl.styles.Font(bold=True)

        # Global values
        ws['A2'] = self.time_from
        ws['B2'] = self.time_until
        ws['D2'] = self.created_timestamp

        # Condition rows
        r = 4
        for cnd in self.conditions:
            ws[f'A{r}'] = cnd.site
            ws[f'B{r}'] = cnd.master_alias
            ws[f'C{r}'] = cnd.condition
            ws[f'D{r}'] = cnd.data_from
            ws[f'E{r}'] = cnd.data_until
            ws[f'F{r}'] = cnd.percentage_valid
            ws[f'G{r}'] = cnd.percentage_notvalid
            ws[f'H{r}'] = cnd.percentage_nodata

            # Percent format
            ws[f'F{r}'].number_format = '0.00 %'
            ws[f'G{r}'].number_format = '0.00 %'
            ws[f'H{r}'].number_format = '0.00 %'

            r += 1

        return wb

    def __getitem__(self, key):
        """
        Returns the Condition instance on the corresponding index.
        """
        return self.conditions[key]

    def __str__(self):
        # TODO: create a meaningful print representation
        out = 'A CondCollection.\n'
        out += '  Time range:\n'
        out += f"    from {self.time_from.strftime('%Y-%m-%d %H:%M:%S')}\n"
        out += f"    to   {self.time_until.strftime('%Y-%m-%d %H:%M:%S')}\n"
        out += f'  {len(self.stations)} stations:\n'
        out += f'    {", ".join(list(self.stations))}\n'
        out += f'  {len(self.conditions)} conditions:\n'
        for c in self.conditions:
            out += f'{str(c)}\n'
        return out

    @classmethod
    def from_dictlist(cls, dictlist, time_from, time_until, pg_conn=None, title=None):
        """
        Create instance and add conditions from list of dicts.
        Dicts must have corresponding keys
        ``'site', 'master_alias', 'raw_condition'``.
        Times must be ``datetime`` objects.
        """
        cc = cls(time_from, time_until, pg_conn, title)
        for d in dictlist:
            cc.add_condition(**d)
        cc.set_sensor_ids()
        return cc

    @classmethod
    def from_xlsx_sheet(cls, ws, pg_conn=None):
        """
        Create a condition collection for analysis
        based on an ``openpyxl`` ``worksheet`` object ``ws``.
        Database connection instance ``pg_conn`` should be prepared
        in advance and passed to this method.

        .. note:: Start and end dates must be in cells A2 and B2, respectively,
                  and conditions must be listed starting from row 4,
                  such that ``site`` is in column A,
                  ``master_alias`` in column B
                  and ``raw_condition`` in column C.
                  There must not be empty rows in between.
                  Any columns outside A:C are ignored,
                  so additional data can be placed outside them.
        """
        # Handle start and end dates;
        # method is interrupted if either one
        # is text BUT is not a valid date
        dateformat = '%d.%m.%Y'
        time_from = ws['A2'].value
        if not isinstance(time_from, datetime):
            time_from = datetime.strptime(ws['A2'].value, dateformat)
        time_until = ws['B2'].value
        if not isinstance(time_until, datetime):
            time_until = datetime.strptime(ws['B2'].value, dateformat)

        # Collect condition rows into a list of dictionaries,
        # make sure the cells have no None values
        dl = []
        for row in ws.iter_rows(min_row=4, max_col=3):
            cells = [c for c in row]
            # Exit upon an empty row
            if not any(cells):
                break
            # TODO: raise an error upon empty cell
            #       or empty row followed by non-empty rows
            dl.append(dict(
                site=cells[0].value,
                master_alias=cells[1].value,
                raw_condition=cells[2].value,
                excel_row=cells[0].row
            ))

        # Now the dictlist method can be used to
        # construct the CondCollection
        cc = cls.from_dictlist(
            dictlist=dl,
            time_from=time_from,
            time_until=time_until,
            pg_conn=pg_conn,
            title=ws.title
        )

        return cc
