#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import os
from io import BytesIO
from flask import Flask, flash, request, redirect, render_template
from flask_bootstrap import Bootstrap
from forms import UploadForm
from werkzeug.utils import secure_filename
from openpyxl import load_workbook

def allowed_file(filename):
	return filename.endswith('.xlsx')

app = Flask(__name__)
Bootstrap(app)
app.secret_key = 'many random bytes'

# @app.route('/')
# def index():
#     return render_template('index.html')

# Check out:
# https://john.soban.ski/pass-bootstrap-html-attributes-to-flask-wtforms.html

@app.route('/', methods=['GET', 'POST'])
def upload_file():
    form = UploadForm()
    if form.validate_on_submit():
        filename = secure_filename(form.file.data.filename)
        wb = load_workbook(form.file.data)
        wsnames = ', '.join(wb.sheetnames)
        flash(f'File {filename} successfully uploaded with sheets {wsnames}', 'success')
        return redirect('/')
    # if 'inputFile' not in request.files:
    #     flash('No file part', 'warning')
    #     return redirect(request.url)
    # file = request.files['inputFile']
    # if file and allowed_file(file.filename):
    #     filename = secure_filename(file.filename)
    #     flash(f'File {filename} successfully uploaded', 'success')
    #     print('Success')
    #     return redirect('/')
    # flash('Oops', 'warning')
    return render_template('index.html', form=form)

app.config['DEBUG'] = True
if __name__ == '__main__':
    app.run()
